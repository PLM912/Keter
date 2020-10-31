import discord
import time
import psutil
import os
import asyncio
import openpyxl
import random
import math
import numpy as np
import datetime
import matplotlib.pyplot as plt

from datetime import datetime
from discord.ext import commands
from evs import default, permissions

shindanlib = "./lib/shindan/"


class Shindan_ko(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.config = default.get("config.json")
        # 폴더생성
        if os.path.isdir("./lib/shindan"):
            print("shindan folder exist")
        else:
            os.makedirs("./lib/shindan")
        if os.path.isfile("./lib/cache/shindan_request.ccf"):
            print("shindan cache file exist")
        else:
            f = open("./lib/cache/shindan_request.ccf", "w")
            f.close()
            f = open("./lib/cache/shindan_requestid.ccf", "w")
            f.close()

    # Commands
    @commands.command(aliases=["진단만들기", "진단생성"])
    async def _say(self, ctx, *, content: str):
        embed = discord.Embed(title="진단메이커", description=content + "에 대한 진단을 요청하시겠습니까?", color=0xeff0f1)
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            if os.path.isfile(shindanlib + content + ".xlsx"):
                await msg.delete()
                embed = discord.Embed(title="진단메이커", description="이미 존재하는 진단입니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                return await ctx.send(embed=embed)
            f = open("./lib/cache/shindan_request.ccf", "a")
            f.write(f"{content}/")
            f.close()
            f = open("./lib/cache/shindan_requestid.ccf", "a")
            f.write(f"{str(ctx.author.id)}/")
            f.close()
            await ctx.send(content + "에 대한 진단을 관리자에게 요청하였습니다.")

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command(aliases=['진단목록', '진단리스트'])
    async def shinlis(self, ctx, plist: int):
        shins = os.listdir(shindanlib)
        embed = discord.Embed(title="진단메이커", color=0xeff0f1)
        for i in range(0 + 10 * (plist - 1), 10 + 10 * (plist - 1)):
            try:
                embed.add_field(name=str(i + 1), value=shins[i].replace(".xlsx", ""))
            except IndexError:
                return await ctx.send(embed=embed)
        await ctx.send(embed=embed)

    @commands.command(aliases=["진단요청", "진단요청목록"])
    async def shinreq(self, ctx):
        f = open("./lib/cache/shindan_request.ccf", "r")
        reqs = f.read().split("/")
        del reqs[-1]
        f.close()
        f = open("./lib/cache/shindan_requestid.ccf", "r")
        reqid = f.read().split("/")
        del reqid[-1]
        f.close()
        embed = discord.Embed(title="진단메이커", description=f"현재 총 {str(len(reqs))}개의 진단 생성 요청이 있습니다.", color=0xeff0f1)
        for i in range(0, 20):
            try:
                embed.add_field(name=f"{str(i)} : {reqs[i]}", value=f"requester : {reqid[i]}", inline=False)
            except:
                pass
        await ctx.send(embed=embed)

    @commands.command(aliases=["진단승락", "진단승인", "진단허가"])
    @commands.check(permissions.is_owner)
    async def shinacs(self, ctx, position: int):
        if position < 1:
            return await ctx.send("**position** 변수는 자연수여야 합니다.")
        f = open("./lib/cache/shindan_request.ccf", "r")
        allreq = f.read()
        reqs = allreq.split("/")
        f.close()
        f = open("./lib/cache/shindan_requestid.ccf", "r")
        allreqid = f.read()
        reqid = allreqid.split("/")
        f.close()
        embed = discord.Embed(title="진단메이커", description=reqs[position - 1] + "에 대한 진단을 만드시겠습니까?", color=0xeff0f1)
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            f = open("./lib/cache/shindan_request.ccf", "w")
            f.write(allreq.replace(f"{reqs[position - 1]}/",""))
            f.close()
            f = open("./lib/cache/shindan_requestid.ccf", "w")
            f.write(allreqid.replace(f"{reqid[position - 1]}/",""))
            f.close()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = str(reqid[position - 1])  #author
            ws.cell(row=1, column=2).value = "진단 <변수1>"  #form
            ws.cell(row=1, column=3).value = "변수1" #vals name
            ws.cell(row=2, column=3).value = "1" #vals count
            for i in range(4, 17):
                ws.cell(row=2, column=3).value = "0"
            ws.cell(row=3, column=1).value = "데이터를 추가해주세요" #vals
            wb.save(shindanlib + f"{reqs[position - 1]}.xlsx")
            wb.close()
            time.sleep(1)
            await msg.delete()
            await ctx.send(reqs[position - 1] + "에 대한 진단을 생성하였습니다.")

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command(aliases=["진단거절"])
    @commands.check(permissions.is_owner)
    async def shinrfs(self, ctx, position: int):
        if position < 1:
            return await ctx.send("**position** 변수는 자연수여야 합니다.")
        f = open("./lib/cache/shindan_request.ccf", "r")
        allreq = f.read()
        reqs = allreq.split("/")
        f.close()
        f = open("./lib/cache/shindan_requestid.ccf", "r")
        allreqid = f.read()
        reqid = allreq.split("/")
        f.close()
        embed = discord.Embed(title="진단메이커", description=reqs[position - 1] + "에 대한 진단요청을 거절하시겠습니까?", color=0xeff0f1)
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            f = open("./lib/cache/shindan_request.ccf", "w")
            f.write(allreq.replace(f"{reqs[position - 1]}/",""))
            f.close()
            f = open("./lib/cache/shindan_requestid.ccf", "w")
            f.write(allreqid.replace(f"{reqid[position - 1]}/",""))
            f.close()
            await msg.delete()
            await ctx.send(reqs[position - 1] + "에 대한 진단요청을 거절하였습니다.")

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command(aliases=["진단초기화"])
    @commands.check(permissions.is_owner)
    async def shinrst(self, ctx):
        f = open("./lib/cache/shindan_request.ccf", "w")
        f.close()
        f = open("./lib/cache/shindan_requestid.ccf", "w")
        f.close()
        await msg.send("진단요청을 초기화하였습니다.")

    @commands.command(aliases=["진단수정"])
    async def shincng(self, ctx, *, shindan: str):
        if os.path.isfile(shindanlib + f"{shindan}.xlsx"):
            wb = openpyxl.load_workbook(shindanlib + f"{shindan}.xlsx")
            ws = wb.active
            if ws.cell(row=1, column=1).value == str(ctx.author.id):
                pass
            else:
                wb.close()
                return await ctx.send("해당 진단을 수정할 수 없습니다.")
        else:
            return await ctx.send("해당 이름의 진단을 찾지 못했습니다.")
        embed = discord.Embed(title="진단메이커", description="진단의 내용을 수정하시겠습니까?", color=0xeff0f1)
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        def reactions_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) in ["ℹ", "🔤", "🆕"]:
                return True
            return False

        def check_(m):
            if m.user_id == ctx.author.id:
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="어떤 내용을 수정하시겠습니까?\nℹ : 내용   🔤 : 변수이름   🆕 : 데이터", color=0xeff0f1)
            msg = await ctx.send(embed=embed)
            await msg.add_reaction("ℹ")
            await msg.add_reaction("🔤")
            await msg.add_reaction("🆕")

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            return await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await msg.edit(content=embed)

        try:
            order = await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reactions_check_)

        except TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="시간이 초과되었습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            return await ctx.send(embed=embed)

        wb = openpyxl.load_workbook(shindanlib + f"{shindan}.xlsx")
        ws = wb.active

        if order == "ℹ":
            msg.delete()
            embed = discord.Embed(title="진단메이커", description="바꿀 내용을 말해주세요.", color=0xeff0f1)
            await msg.send(content=embed)
            try:
                newctx = await self.bot.wait_for('message', timeout=120.0, check=check_)
                ws.cell(row=1, column=1).value = newctx.content
                await newctx.add_reaction("👍")

            except TimeoutError:
                await msg.delete()
                embed = discord.Embed(title="진단메이커", description="시간이 초과되었습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                return await ctx.send(embed=embed)

        if order == "🔤":
            msg.delete()
            embed = discord.Embed(title="진단메이커", description="바꿀 변수의 이름를 말해주세요.", color=0xeff0f1)
            await msg.send(content=embed)
            try:
                newval = await self.bot.wait_for('message', timeout=60.0, check=check_)
                try:
                    position = int(newval.split(" ")[0])
                except:
                    embed = discord.Embed(title="진단메이커", description="변수가 잘못되었습니다.", color=0xeff0f1)
                    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                    return await ctx.send(embed=embed)
                if position > 16:
                    embed = discord.Embed(title="진단메이커", description="변수는 최대 16개까지만 지원합니다.", color=0xeff0f1)
                    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                    return await ctx.send(embed=embed)
                if position < 1:
                    embed = discord.Embed(title="진단메이커", description="변수가 잘못되었습니다.", color=0xeff0f1)
                    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                    return await ctx.send(embed=embed)
                ws.cell(row=1, column=position).value = newval.content.replace(str(position) + " ", "")
                await newctx.add_reaction("👍")

            except TimeoutError:
                await msg.delete()
                embed = discord.Embed(title="진단메이커", description="시간이 초과되었습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                return await ctx.send(embed=embed)

        if order == "🆕":
            msg.delete()
            embed = discord.Embed(title="진단메이커", description="추가할 데이터를 말해주세요.", color=0xeff0f1)
            await msg.send(content=embed)
            try:
                newval = await self.bot.wait_for('message', timeout=60.0, check=check_)
                try:
                    position = int(newval.split(" ")[0])
                except:
                    embed = discord.Embed(title="진단메이커", description="변수가 잘못되었습니다.", color=0xeff0f1)
                    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                    return await ctx.send(embed=embed)
                if position > 16:
                    embed = discord.Embed(title="진단메이커", description="변수는 최대 16개까지만 지원합니다.", color=0xeff0f1)
                    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                    return await ctx.send(embed=embed)
                if position < 1:
                    embed = discord.Embed(title="진단메이커", description="변수가 잘못되었습니다.", color=0xeff0f1)
                    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                    return await ctx.send(embed=embed)
                ws.cell(row=2, column=position + 2).value = str(int(ws.cell(row=2, column=position + 2).value) + 1)
                ws.cell(row=position, column=int(ws.cell(row=2, column=position).value) + 1).value = newval.content.replace(str(position) + " ", "")
                await newctx.add_reaction("👍")

            except TimeoutError:
                await msg.delete()
                embed = discord.Embed(title="진단메이커", description="시간이 초과되었습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                return await ctx.send(embed=embed)

        wb.close()

    @commands.command(aliases=["진단정보"])
    async def shininf(self, ctx,  *, shindan: str):
        if os.path.isfile(shindanlib + f"{shindan}.xlsx"):
            wb = openpyxl.load_workbook(shindanlib + f"{shindan}.xlsx")
            ws = wb.active
            authorid = ws.cell(row=1, column=1).value
            cases = 1
            for i in range(3, 20):
                try:
                    if int(ws.cell(row=2, column=i).value) == 0:
                        pass
                    else:
                        cases = cases * int(ws.cell(row=2, column=i).value)
                except:
                    pass
            embed = discord.Embed(title="진단메이커", description=f"{shindan}의 정보", color=0xeff0f1)
            embed.add_field(name="제작자", value=authorid)
            embed.add_field(name="경우의 수", value=cases)
            wb.close()
            return await ctx.send(embed=embed)
        await msg.delete()
        embed = discord.Embed(title="진단메이커", description="해당 이름의 진단을 찾을 수 없습니다.", color=0xeff0f1)
        embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
        return await ctx.send(embed=embed)

    @commands.command(aliases=["진단삭제"])
    @commands.check(permissions.is_owner)
    async def shindel(self, ctx, shindan: str):
        if not os.path.isfile(shindanlib + f"{shindan}.xlsx"):
            embed = discord.Embed(title="진단메이커", description="해당 이름의 진단을 찾을 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            return await ctx.send(embed=embed)
        wb = openpyxl.load_workbook(shindanlib + f"{shindan}.xlsx")
        ws = wb.active
        authorid = str(ws.cell(row=1, column=1).value)
        wb.close()
        
        if not authorid == str(ctx.author.id):
            embed = discord.Embed(title="진단메이커", description="해당 진단을 삭제할 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            return await ctx.send(embed=embed)

        embed = discord.Embed(title="진단메이커", description=shindan + "에 대한 진단을 삭제하시겠습니까?\n 해당 진단은 영구적으로 삭제될 것입니다.", color=0xeff0f1)
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            await msg.delete()
            os.remove(shindanlib + shindan + ".xlsx")
            await ctx.send(shindan + "에 대한 진단을 삭제하였습니다.")

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command(aliases=["관리자진단삭제", "관진삭"])
    @commands.check(permissions.is_owner)
    async def shindel(self, ctx, shindan: str):
        if not os.path.isfile(shindanlib + f"{shindan}.xlsx"):
            embed = discord.Embed(title="진단메이커", description="해당 이름의 진단을 찾을 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            return await ctx.send(embed=embed)
        wb = openpyxl.load_workbook(shindanlib + f"{shindan}.xlsx")
        ws = wb.active
        authorid = str(ws.cell(row=1, column=1).value)
        wb.close()

        embed = discord.Embed(title="진단메이커", description=shindan + "에 대한 진단을 삭제하시겠습니까?\n 해당 진단은 영구적으로 삭제될 것입니다.", color=0xeff0f1)
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            await msg.delete()
            os.remove(shindanlib + shindan + ".xlsx")
            await ctx.send(shindan + "에 대한 진단을 삭제처리하였습니다.")

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="진단메이커", description="동의하지 않으셨습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)


def setup(bot):
    bot.add_cog(Shindan_ko(bot))
